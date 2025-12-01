#!/usr/bin/env python3
"""
Import Manager for Dram Planner

Handles importing bottles from various formats: CSV, JSON, Excel
"""

import json
import csv
import os
from datetime import datetime

# Try to import optional dependencies
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def validate_bottle_data(bottle_data, row_num=None):
    """
    Validate bottle data structure and values.
    
    Args:
        bottle_data (dict): Bottle data to validate.
        row_num (int, optional): Row number for error reporting.
        
    Returns:
        tuple: (is_valid, errors_list)
    """
    errors = []
    
    # Required fields
    if 'name' not in bottle_data or not bottle_data.get('name', '').strip():
        errors.append(f"Row {row_num}: Missing or empty 'name' field")
    
    if 'category' not in bottle_data or not bottle_data.get('category', '').strip():
        errors.append(f"Row {row_num}: Missing or empty 'category' field")
    
    # Validate ABV if provided
    if 'abv' in bottle_data and bottle_data['abv'] is not None:
        try:
            abv = float(bottle_data['abv'])
            if abv < 0 or abv > 100:
                errors.append(f"Row {row_num}: ABV {abv}% is out of valid range (0-100)")
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: Invalid ABV value: {bottle_data['abv']}")
    
    # Validate price if provided
    if 'price_paid' in bottle_data and bottle_data['price_paid'] is not None:
        try:
            price = float(bottle_data['price_paid'])
            if price < 0:
                errors.append(f"Row {row_num}: Price cannot be negative: {price}")
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: Invalid price value: {bottle_data['price_paid']}")
    
    # Validate date format if provided
    if 'purchase_date' in bottle_data and bottle_data.get('purchase_date'):
        try:
            datetime.strptime(bottle_data['purchase_date'], '%Y-%m-%d')
        except ValueError:
            errors.append(f"Row {row_num}: Invalid date format '{bottle_data['purchase_date']}'. Use YYYY-MM-DD")
    
    return len(errors) == 0, errors


def normalize_bottle_data(bottle_data):
    """
    Normalize bottle data to standard format.
    
    Args:
        bottle_data (dict): Raw bottle data.
        
    Returns:
        dict: Normalized bottle data.
    """
    normalized = {
        'name': str(bottle_data.get('name', '')).strip(),
        'category': str(bottle_data.get('category', 'other')).strip().lower(),
        'abv': _safe_float(bottle_data.get('abv'), 0.0),
        'price_paid': _safe_float(bottle_data.get('price_paid'), 0.0),
        'purchase_date': str(bottle_data.get('purchase_date', '')).strip(),
        'opened_date': str(bottle_data.get('opened_date', '')).strip(),
        'notes': str(bottle_data.get('notes', '')).strip(),
        'barcode': str(bottle_data.get('barcode', '')).strip(),
        'tasted': bool(bottle_data.get('tasted', False)),
        'tasting_date': bottle_data.get('tasting_date'),
        'rating': _safe_float(bottle_data.get('rating')),
        'tasting_notes': str(bottle_data.get('tasting_notes', '')).strip()
    }
    
    return normalized


def _safe_float(value, default=None):
    """Safely convert value to float."""
    if value is None or value == '':
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def import_from_csv(csv_file, preview=False):
    """
    Import bottles from CSV file.
    
    Supports multiple CSV formats:
    - name,category,abv,price_paid,purchase_date,notes,barcode
    - Header row is optional and auto-detected
    
    Args:
        csv_file (str): Path to CSV file.
        preview (bool): If True, return preview without validation errors.
        
    Returns:
        tuple: (bottles_list, errors_list, warnings_list)
    """
    bottles = []
    errors = []
    warnings = []
    
    if not os.path.exists(csv_file):
        errors.append(f"CSV file not found: {csv_file}")
        return bottles, errors, warnings
    
    try:
        with open(csv_file, 'r', encoding='utf-8') as f:
            # Try to detect delimiter
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            try:
                delimiter = sniffer.sniff(sample).delimiter
            except:
                delimiter = ','
            
            reader = csv.DictReader(f, delimiter=delimiter)
            
            # If no header, use positional columns
            if not reader.fieldnames or len(reader.fieldnames) == 0:
                f.seek(0)
                reader = csv.reader(f)
                header_skipped = False
                row_num = 0
                
                for row in reader:
                    row_num += 1
                    if not header_skipped and row and row[0].lower() in ['name', 'bottle', 'spirit']:
                        header_skipped = True
                        continue
                    
                    if len(row) < 2:
                        warnings.append(f"Row {row_num}: Skipped (insufficient data)")
                        continue
                    
                    bottle_data = {
                        'name': row[0] if len(row) > 0 else '',
                        'category': row[1] if len(row) > 1 else 'other',
                        'abv': row[2] if len(row) > 2 else None,
                        'price_paid': row[3] if len(row) > 3 else None,
                        'purchase_date': row[4] if len(row) > 4 else None,
                        'notes': row[5] if len(row) > 5 else None,
                        'barcode': row[6] if len(row) > 6 else None
                    }
                    
                    normalized = normalize_bottle_data(bottle_data)
                    is_valid, validation_errors = validate_bottle_data(normalized, row_num)
                    
                    if is_valid:
                        bottles.append(normalized)
                    else:
                        errors.extend(validation_errors)
            else:
                # Use DictReader with headers
                row_num = 1  # Start at 1 (header is row 0)
                for row in reader:
                    row_num += 1
                    normalized = normalize_bottle_data(row)
                    is_valid, validation_errors = validate_bottle_data(normalized, row_num)
                    
                    if is_valid:
                        bottles.append(normalized)
                    else:
                        errors.extend(validation_errors)
                        
    except UnicodeDecodeError:
        errors.append(f"Error: Could not decode CSV file. Try UTF-8 encoding.")
    except Exception as e:
        errors.append(f"Error reading CSV file: {e}")
    
    return bottles, errors, warnings


def import_from_json(json_file, preview=False):
    """
    Import bottles from JSON file.
    
    Supports formats:
    - Array of bottle objects: [{"name": "...", "category": "..."}, ...]
    - Object with 'bottles' key: {"bottles": [{"name": "...", ...}, ...]}
    
    Args:
        json_file (str): Path to JSON file.
        preview (bool): If True, return preview without validation errors.
        
    Returns:
        tuple: (bottles_list, errors_list, warnings_list)
    """
    bottles = []
    errors = []
    warnings = []
    
    if not os.path.exists(json_file):
        errors.append(f"JSON file not found: {json_file}")
        return bottles, errors, warnings
    
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Extract bottles array
        if isinstance(data, list):
            bottles_data = data
        elif isinstance(data, dict):
            if 'bottles' in data:
                bottles_data = data['bottles']
            else:
                errors.append("JSON file must contain 'bottles' array or be an array of bottles")
                return bottles, errors, warnings
        else:
            errors.append("JSON file must be an object or array")
            return bottles, errors, warnings
        
        if not isinstance(bottles_data, list):
            errors.append("Bottles data must be an array")
            return bottles, errors, warnings
        
        # Process each bottle
        for idx, bottle_data in enumerate(bottles_data, start=1):
            if not isinstance(bottle_data, dict):
                errors.append(f"Bottle {idx}: Must be an object")
                continue
            
            normalized = normalize_bottle_data(bottle_data)
            is_valid, validation_errors = validate_bottle_data(normalized, idx)
            
            if is_valid:
                bottles.append(normalized)
            else:
                errors.extend(validation_errors)
                
    except json.JSONDecodeError as e:
        errors.append(f"Invalid JSON: {e}")
    except Exception as e:
        errors.append(f"Error reading JSON file: {e}")
    
    return bottles, errors, warnings


def import_from_excel(excel_file, sheet_name=None, preview=False):
    """
    Import bottles from Excel file (.xlsx).
    
    Args:
        excel_file (str): Path to Excel file.
        sheet_name (str, optional): Sheet name. Uses first sheet if not specified.
        preview (bool): If True, return preview without validation errors.
        
    Returns:
        tuple: (bottles_list, errors_list, warnings_list)
    """
    bottles = []
    errors = []
    warnings = []
    
    if not EXCEL_AVAILABLE:
        errors.append("Excel import requires 'openpyxl' library. Install with: pip install openpyxl")
        return bottles, errors, warnings
    
    if not os.path.exists(excel_file):
        errors.append(f"Excel file not found: {excel_file}")
        return bottles, errors, warnings
    
    try:
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Select sheet
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                errors.append(f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(workbook.sheetnames)}")
                return bottles, errors, warnings
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        # Read header row
        headers = []
        header_row = sheet[1]
        for cell in header_row:
            headers.append(str(cell.value).lower().strip() if cell.value else '')
        
        # Map common header variations
        header_map = {
            'name': ['name', 'bottle', 'spirit', 'product', 'title'],
            'category': ['category', 'type', 'kind'],
            'abv': ['abv', 'alcohol', 'alcohol %', 'proof'],
            'price_paid': ['price', 'price_paid', 'cost', 'purchase price'],
            'purchase_date': ['purchase_date', 'date', 'purchase date', 'bought'],
            'notes': ['notes', 'note', 'description', 'desc'],
            'barcode': ['barcode', 'upc', 'ean', 'code']
        }
        
        # Find column indices
        column_indices = {}
        for key, variations in header_map.items():
            for idx, header in enumerate(headers):
                if header in variations:
                    column_indices[key] = idx
                    break
        
        # Read data rows
        row_num = 1
        for row in sheet.iter_rows(min_row=2, values_only=False):
            row_num += 1
            row_values = [cell.value for cell in row]
            
            # Skip empty rows
            if not any(row_values):
                continue
            
            # Build bottle data from columns
            bottle_data = {}
            for key, idx in column_indices.items():
                if idx < len(row_values) and row_values[idx] is not None:
                    bottle_data[key] = row_values[idx]
            
            # Fallback to positional if no headers found
            if not column_indices:
                if len(row_values) >= 2:
                    bottle_data = {
                        'name': row_values[0],
                        'category': row_values[1],
                        'abv': row_values[2] if len(row_values) > 2 else None,
                        'price_paid': row_values[3] if len(row_values) > 3 else None,
                        'purchase_date': row_values[4] if len(row_values) > 4 else None,
                        'notes': row_values[5] if len(row_values) > 5 else None,
                        'barcode': row_values[6] if len(row_values) > 6 else None
                    }
                else:
                    warnings.append(f"Row {row_num}: Skipped (insufficient data)")
                    continue
            
            normalized = normalize_bottle_data(bottle_data)
            is_valid, validation_errors = validate_bottle_data(normalized, row_num)
            
            if is_valid:
                bottles.append(normalized)
            else:
                errors.extend(validation_errors)
                
    except Exception as e:
        errors.append(f"Error reading Excel file: {e}")
    
    return bottles, errors, warnings


def preview_import(import_func, *args, **kwargs):
    """
    Preview import without committing.
    
    Args:
        import_func: Import function to call.
        *args, **kwargs: Arguments to pass to import function.
        
    Returns:
        tuple: (bottles_list, errors_list, warnings_list)
    """
    kwargs['preview'] = True
    return import_func(*args, **kwargs)


def print_import_preview(bottles, errors, warnings):
    """Print a preview of the import."""
    print(f"\n{'='*60}")
    print(f"Import Preview")
    print(f"{'='*60}")
    print(f"Valid bottles: {len(bottles)}")
    print(f"Errors: {len(errors)}")
    print(f"Warnings: {len(warnings)}")
    
    if warnings:
        print(f"\nWarnings:")
        for warning in warnings[:10]:
            print(f"  - {warning}")
        if len(warnings) > 10:
            print(f"  ... and {len(warnings) - 10} more")
    
    if errors:
        print(f"\nErrors:")
        for error in errors[:10]:
            print(f"  - {error}")
        if len(errors) > 10:
            print(f"  ... and {len(errors) - 10} more")
    
    if bottles:
        print(f"\nSample bottles (first 5):")
        for i, bottle in enumerate(bottles[:5], 1):
            print(f"  {i}. {bottle['name']} ({bottle['category']}) - ABV: {bottle.get('abv', 'N/A')}%")
        if len(bottles) > 5:
            print(f"  ... and {len(bottles) - 5} more")

